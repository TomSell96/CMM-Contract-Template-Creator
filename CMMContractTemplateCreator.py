import openpyxl as OPXL
from openpyxl import Workbook
from openpyxl import load_workbook

def addFirstSupplier():

    contSchedule = input("Please enter the file name of the contract schedule and the file extension e.g. 42583.xlsx: ")

    blankFile = input("Please enter the file name of the file to write the data to and the file extension e.g. AccordOralLiquid.xlsx: ")

    readPath = ""+ contSchedule

    writePath = ""+ blankFile

    supplierPath =  ""

    comments = "Added via python script"

    #load the file
    contractNWSSP = load_workbook(readPath)

    cmmTemplate = load_workbook(writePath)

    #create active worksheet

    wsNWSSP = contractNWSSP.active

    wsCMM = cmmTemplate.active

    #Defining Other Variables

    print("Please enter contract name")
    contractName = input()

    print("Please enter contract description")
    contractDescription = input()

    print("Please enter contract number")
    contractNumber = input()

    print("Please enter contract start date")
    startDate = input()

    print("Please enter contract end date")
    endDate = input()

    print("What supplier from the contract is required? Please write as per spreadsheet")
    contractSupplier = input()

    print("What is the suppliers name on WHPSMS")
    contMainSupplier = input()

    print("What wholesalers are required? \nPlease enter one at a time and select enter \nFor AAH enter AAH \nFor Alliance enter ALLI \nFor Mawdsleys enter MAW \nFor Phoenix enter PHOE \nFor Alloga enter ALLOGA \nFor PolarSpeed enter POLAR \nFor Lloyds Pharmacy Clinical Homecare enter LLOYDS \nFor Sciencus enter SCIEN \nFor Healthnet enter HEALTHNET \nFor Pharmaxo enter PHARMAXO \nFor Alcura enter ALCURA \nPlease enter BLANK for non-required wholesalers")

    wholesaler1 = input().upper()

    wholesaler2 = input().upper()

    wholesaler3 = input().upper()

    wholesaler4 = input().upper()

    wholesaler5 = input().upper()

    wholesalerInput = {
        
        "AAH":"3333/S;3333/B;3333/W",
        "ALLI":"30228/S;30228/P",
        "MAW":"181192/S;181192/MK",
        "PHOE": "216686/W;216686/C",
        "ALLOGA": "30459",
        "POLAR": "167912341",
        "LLOYDS": "34874567543",
        "SCIEN": "3292345563",
        "HEALTHNET": "63035890605",
        "ALCURA": "550234788",
        "PHARMAXO": "511356911"
    }

    for i, j in wholesalerInput.items():
        if wholesaler1 == i:
            wholesaler1input = j
        

        if wholesaler2 == i:
            wholesaler2input = j
        

        if wholesaler3 == i:
            wholesaler3input = j
            

        if wholesaler4 == i:
            wholesaler4input = j
            

        if wholesaler5 == i:
            wholesaler5input = j
            

    if wholesaler1 == "BLANK" and wholesaler2 == "BLANK" and wholesaler3 == "BLANK" and wholesaler4 == "BLANK" and wholesaler5 == "BLANK":
        contractwholesaler = wholesaler1input

    elif wholesaler1 != "BLANK" and wholesaler2 == "BLANK" and wholesaler3 == "BLANK" and wholesaler4 == "BLANK" and wholesaler5 == "BLANK":
        contractwholesaler = wholesaler1input

    elif wholesaler1 != "BLANK" and wholesaler2 != "BLANK" and wholesaler3 == "BLANK" and wholesaler4 == "BLANK" and wholesaler5 == "BLANK":
        contractwholesaler = wholesaler1input + ";" + wholesaler2input

    elif wholesaler1 != "BLANK" and wholesaler2 != "BLANK" and wholesaler3 != "BLANK" and wholesaler4 == "BLANK" and wholesaler5 == "BLANK":
        contractwholesaler = wholesaler1input + ";" + wholesaler2input + ";" + wholesaler3input

    elif wholesaler1 != "BLANK" and wholesaler2 != "BLANK" and wholesaler3 != "BLANK" and wholesaler4 != "BLANK" and wholesaler5 == "BLANK":
        contractwholesaler = wholesaler1input + ";" + wholesaler2input + ";" + wholesaler3input + ";" + wholesaler4input

    elif wholesaler1 != "BLANK" and wholesaler2 != "BLANK" and wholesaler3 != "BLANK" and wholesaler4 != "BLANK" and wholesaler5 != "BLANK":
        contractwholesaler = wholesaler1input + ";" + wholesaler2input + ";" + wholesaler3input + ";" + wholesaler4input + ";" + wholesaler5input

    contractStatus = "Draft"

    supplierWSK = load_workbook(supplierPath)

    wsSupplierWSK = supplierWSK.active

    # Define suppliers dictionary inside the for loop
    for suppliers in wsSupplierWSK.iter_rows():
        suppliers = {suppliers[0].value: suppliers[1].value}

        for supplier, supplierNumber in suppliers.items():
            if contMainSupplier == supplier:
                contMainSupplierInput = supplierNumber

    print("Grabbing the data from spreadsheet")

    contList = []

    #For loop to grab the values out of the columns and assign to the corresponding list variable
    for rows in wsNWSSP.iter_rows():
        drug = {
            "Contract Name": contractName,
            "Contract Description": contractDescription,
            "Contract Number": contractNumber,
            "Start Date": startDate,
            "End Date": endDate,
            "Comments": comments,
            "Status": "Draft",
            "Main Supplier": rows[0].value,
            "Honouring Suppliers": contractwholesaler,
            "Drug + Pack": rows[7].value,
            "Minimum Order Quantity": 1,
            "Minimum Order Quantity Units": 1,
            "Drug Comments": "Added via python script",
            "Discount - Quantity": 1,
            "Discount - Price/Unit": rows[5].value,
            "Discount - %": ""

            }
        contList.append(drug)


    selectContractSuppliers = list(filter(lambda drug: drug['Main Supplier'] == contractSupplier, contList))

    for item in selectContractSuppliers:
        item.update((k, contMainSupplierInput) for k, v in item.items() if v == contractSupplier)

    wb = load_workbook(writePath)
    ws = wb.active
    wb.save(writePath)
    
    fieldNames = ["Contract Name" , "Contract Description", "Contract Number" , "Start Date", "End Date" , "Comments" , "Status" , "Main Supplier" , "Honouring Suppliers", "Drug + Pack" , "Minimum Order Quantity" , "Minimum Order Quantity Units" , "Drug Comments" , "Discount - Quantity" , "Discount - Price/Unit" , "Discount - %"]
    
    ws.append(fieldNames)
    
    for drugs in selectContractSuppliers:
        contractLines = [drugs[k] for k in fieldNames]
        ws.append(contractLines)
        wb.save(writePath)
    
    for row_values in ws.iter_rows(values_only=True):
        for value in row_values:
            wb.save(writePath)
def addAnotherSupplierFunc():
    contSchedule = input("Please enter the file name of the contract schedule and the file extension e.g. 42583.xlsx: ")

    blankFile = input("Please enter the file name of the file to write the data to and the file extension e.g. AccordOralLiquid.xlsx: ")

    readPath = "C:/Users/th252839/Desktop/Pandas/"+ contSchedule

    writePath = "C:/Users/th252839/Desktop/Pandas/"+ blankFile

    supplierPath =  "C:/Users/th252839/Desktop/Pandas/supplierName.xlsx"

    comments = "Added via python script"

    #load the file
    contractNWSSP = load_workbook(readPath)

    cmmTemplate = load_workbook(writePath)

    #create active worksheet

    wsNWSSP = contractNWSSP.active

    wsCMM = cmmTemplate.active

    #Defining Other Variables

    print("Please enter contract name")
    contractName = input()

    print("Please enter contract description")
    contractDescription = input()

    print("Please enter contract number")
    contractNumber = input()

    print("Please enter contract start date")
    startDate = input()

    print("Please enter contract end date")
    endDate = input()

    print("What supplier from the contract is required? Please write as per spreadsheet")
    contractSupplier = input()

    print("What is the suppliers name on WHPSMS")
    contMainSupplier = input()

    print("What wholesalers are required? \nPlease enter one at a time and select enter \nFor AAH enter AAH \nFor Alliance enter ALLI \nFor Mawdsleys enter MAW \nFor Phoenix enter PHOE \nFor Alloga enter ALLOGA \nFor PolarSpeed enter POLAR \nFor Lloyds Pharmacy Clinical Homecare enter LLOYDS \nFor Sciencus enter SCIEN \nFor Healthnet enter HEALTHNET \nFor Pharmaxo enter PHARMAXO \nFor Alcura enter ALCURA \nPlease enter BLANK for non-required wholesalers")

    wholesaler1 = input().upper()

    wholesaler2 = input().upper()

    wholesaler3 = input().upper()

    wholesaler4 = input().upper()

    wholesaler5 = input().upper()

    wholesalerInput = {
        
        "AAH":"3/S;3/B;3/W",
        "ALLI":"308/S;308/P",
        "MAW":"18192/S;18192/MK",
        "PHOE": "21686/W;21686/C",
        "ALLOGA": "309",
        "POLAR": "16791",
        "LLOYDS": "34873",
        "SCIEN": "3293",
        "HEALTHNET": "63005",
        "ALCURA": "5508",
        "PHARMAXO": "51911"
    }

    for i, j in wholesalerInput.items():
        if wholesaler1 == i:
            wholesaler1input = j
        

        if wholesaler2 == i:
            wholesaler2input = j
        

        if wholesaler3 == i:
            wholesaler3input = j
            

        if wholesaler4 == i:
            wholesaler4input = j
            

        if wholesaler5 == i:
            wholesaler5input = j
            

    if wholesaler1 == "BLANK" and wholesaler2 == "BLANK" and wholesaler3 == "BLANK" and wholesaler4 == "BLANK" and wholesaler5 == "BLANK":
        contractwholesaler = wholesaler1input

    elif wholesaler1 != "BLANK" and wholesaler2 == "BLANK" and wholesaler3 == "BLANK" and wholesaler4 == "BLANK" and wholesaler5 == "BLANK":
        contractwholesaler = wholesaler1input

    elif wholesaler1 != "BLANK" and wholesaler2 != "BLANK" and wholesaler3 == "BLANK" and wholesaler4 == "BLANK" and wholesaler5 == "BLANK":
        contractwholesaler = wholesaler1input + ";" + wholesaler2input

    elif wholesaler1 != "BLANK" and wholesaler2 != "BLANK" and wholesaler3 != "BLANK" and wholesaler4 == "BLANK" and wholesaler5 == "BLANK":
        contractwholesaler = wholesaler1input + ";" + wholesaler2input + ";" + wholesaler3input

    elif wholesaler1 != "BLANK" and wholesaler2 != "BLANK" and wholesaler3 != "BLANK" and wholesaler4 != "BLANK" and wholesaler5 == "BLANK":
        contractwholesaler = wholesaler1input + ";" + wholesaler2input + ";" + wholesaler3input + ";" + wholesaler4input

    elif wholesaler1 != "BLANK" and wholesaler2 != "BLANK" and wholesaler3 != "BLANK" and wholesaler4 != "BLANK" and wholesaler5 != "BLANK":
        contractwholesaler = wholesaler1input + ";" + wholesaler2input + ";" + wholesaler3input + ";" + wholesaler4input + ";" + wholesaler5input

    contractStatus = "Draft"

    supplierWSK = load_workbook(supplierPath)

    wsSupplierWSK = supplierWSK.active

    # Define suppliers dictionary inside the for loop
    for suppliers in wsSupplierWSK.iter_rows():
        suppliers = {suppliers[0].value: suppliers[1].value}

        for supplier, supplierNumber in suppliers.items():
            if contMainSupplier == supplier:
                contMainSupplierInput = supplierNumber

    print("Grabbing the data from spreadsheet")

    contList = []

    #For loop to grab the values out of the columns and assign to the corresponding list variable
    for rows in wsNWSSP.iter_rows():
        drug = {
            "Contract Name": contractName,
            "Contract Description": contractDescription,
            "Contract Number": contractNumber,
            "Start Date": startDate,
            "End Date": endDate,
            "Comments": comments,
            "Status": "Draft",
            "Main Supplier": rows[0].value,
            "Honouring Suppliers": contractwholesaler,
            "Drug + Pack": rows[7].value,
            "Minimum Order Quantity": 1,
            "Minimum Order Quantity Units": 1,
            "Drug Comments": "Added via python script",
            "Discount - Quantity": 1,
            "Discount - Price/Unit": rows[5].value,
            "Discount - %": ""

            }
        contList.append(drug)


    selectContractSuppliers = list(filter(lambda drug: drug['Main Supplier'] == contractSupplier, contList))

    for item in selectContractSuppliers:
        item.update((k, contMainSupplierInput) for k, v in item.items() if v == contractSupplier)

    wb = load_workbook(writePath)
    ws = wb.active
    wb.save(writePath)
    
    fieldNames = ["Contract Name" , "Contract Description", "Contract Number" , "Start Date", "End Date" , "Comments" , "Status" , "Main Supplier" , "Honouring Suppliers", "Drug + Pack" , "Minimum Order Quantity" , "Minimum Order Quantity Units" , "Drug Comments" , "Discount - Quantity" , "Discount - Price/Unit" , "Discount - %"]

    for drugs in selectContractSuppliers:
        contractLines = [drugs[k] for k in fieldNames]
        ws.append(contractLines)
        wb.save(writePath)
    
    for row_values in ws.iter_rows(values_only=True):
        for value in row_values:
            wb.save(writePath)

print("Hello, this is the contract upload template populator, for additions or issues add these to the CMM Config channel on teams")

addFirstSupplier()

while True:
    addAnothersupplier = input("Would you like to add another supplier? (Type quit to exit): ")
    if addAnothersupplier.lower() == "quit":
        break
    
    else:
        addAnotherSupplierFunc()

